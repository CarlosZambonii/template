import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import io
import requests

# === Função para aplicar estilo ===
def aplicar_formatacao(ws, cell_range, fill_color=None, font_name=None,
                       font_size=None, bold=None, underline=False,
                       font_color=None, align="center", border=False):
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws[cell_range]:
        for cell in row:
            if any([font_name, font_size, bold, underline, font_color]):
                cell.font = Font(
                    name=font_name or cell.font.name,
                    size=font_size or cell.font.size,
                    bold=bold if bold is not None else cell.font.bold,
                    underline="single" if underline else None,
                    color=font_color or cell.font.color
                )
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color,
                                        end_color=fill_color,
                                        fill_type="solid")
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if border:
                cell.border = thin_border
         

# === Streamlit App ===
st.title("Editor de Ficheiros Manteivias")

# Escolher tipo de relatório
opcao = st.radio(
    "Escolha o tipo de relatório:",
    ["Folha de Ponto (Justificação)", "Resumo Mensal Pessoal (Justificação)", "Resumo Mensal Pessoal (Obras)", "Folha de Ponto (Fiscalização) Com AV", "Folha de Ponto (Fiscalização) Sem AV"]
)

# Upload do arquivo Excel
uploaded_file = st.file_uploader("Escolha o arquivo Excel", type=["xlsx"])

if uploaded_file:
    wb = load_workbook(uploaded_file)
    ws = wb.active

    if opcao == "Folha de Ponto (Justificação)":
        # --- Limpar A1 e remover mesclagens ---
        ws["A1"].value = None 
        
        # --- Remove colunas F e G ---
        ws.delete_cols(6, 2)

        # --- Adicionar logo via URL ---
        logo_url = "https://storage.googleapis.com/hostinger-horizons-assets-prod/3f69205a-c8f5-4046-89f2-58ab4070596a/a445b8a745a1d79fa57a195f52fac4b4.png"
        response = requests.get(logo_url)
        img_bytes = io.BytesIO(response.content)
        img = Image(img_bytes)
        img.height = 40
        img.width = 100
        
        ws.add_image(img, "A1")

        # --- Título ---
        
        ws["A1"].value = None
        ws.unmerge_cells("A1:U1")
        ws.merge_cells("A1:B2")
        ws.merge_cells("AK1:AK2")
        ws.merge_cells("AL1:AL2")
        aplicar_formatacao(ws, "A1:B2", fill_color="D8D8D8")
        ws["C1"].value = "CONTABILIDADE MANTEIVIAS JUSTIFICAÇÕES"
        ws.unmerge_cells("X1:AL1")
        ws.merge_cells("C1:AJ2")
        aplicar_formatacao(ws, "C1:AL2", font_size=12, bold=True, underline=True, fill_color="D8D8D8")
        
        
        # --- Linha 3 ---
        aplicar_formatacao(ws, "A3:AL3", font_size=9, bold=True,
                           font_color="0070C0", fill_color="F2F2F2")

        # --- Coluna A ---
        aplicar_formatacao(ws, "A4:A120", font_size=9, bold=True,
                           font_color="0070C0", fill_color="D8D8D8")

        # --- Coluna B ---
        aplicar_formatacao(ws, "B4:B120", font_size=9, bold=True,
                           font_color="000000")

        # --- Bordas ---
        aplicar_formatacao(ws, "A1:AL120", border=True)

        # --- Pintar células com siglas ---
        cores_siglas = {
            "AT": "C46C00", "Bx": "B7B7E8", "FP": "FF0000", "FE": "FFC0C0",
            "FJ": "D99600", "FI": "00B050", "N": "00B0F0", "S": "7030A0",
            "E": "FFFF00", "LP": "00B050", "LC": "B7B7E8", "I.P.": "808000",
            "F": "FF9900", "FM": "002060", "EE": "C0C0A0"
        }
        for row in ws["A4:AL120"]:
            for cell in row:
                valor = str(cell.value).strip() if cell.value else ""
                if valor in cores_siglas:
                    cor = cores_siglas[valor]
                    cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                    cell.font = Font(bold=True, color="000000")

    elif opcao == "Resumo Mensal Pessoal (Justificação)":
       
        ws["A1"].value = "Resumo Mensal"
        aplicar_formatacao(ws, "A1:F1", font_size=12, bold=True, fill_color="B8CCE4")
         #limites na planilha toda

        aplicar_formatacao(ws, "A1:W117", border=True)

        #Coluna A

        aplicar_formatacao(ws, "A1:A117", font_size=12, fill_color="8DB3E2")

        #justificaçoes
        aplicar_formatacao(ws, "G1:G1", font_size=12, fill_color="8DB3E2", bold=True)
        aplicar_formatacao(ws, "H1:H117", font_size=12, fill_color="D99594", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "I1:I117", font_size=12, fill_color="B2A1C7", font_color="000000", bold=True)
        aplicar_formatacao(ws, "J1:J117", font_size=12, fill_color="00B050", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "K1:K117", font_size=12, fill_color="92D050", font_color="000000", bold=True)
        aplicar_formatacao(ws, "L1:L117", font_size=12, fill_color="E36C09", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "M1:M117", font_size=12, fill_color="C46C00", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "N1:N117", font_size=12, fill_color="FF0000", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "O1:O117", font_size=12, fill_color="00B0F0", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "P1:P117", font_size=12, fill_color="7030A0", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "Q1:Q117", font_size=12, fill_color="FFFF00", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "R1:R117", font_size=12, fill_color="B7B7E8", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "S1:S117", font_size=12, fill_color="808000", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "T1:T117", font_size=12, fill_color="FF0000", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "U1:U117", font_size=12, fill_color="FF9900", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "V1:V117", font_size=12, fill_color="002060", font_color="FFFFFF", bold=True)
        aplicar_formatacao(ws, "W1:W117", font_size=12, fill_color="C0C0A0", font_color="FFFFFF", bold=True)

    elif opcao == "Resumo Mensal Pessoal (Obras)": 
        ws["A1"].value = "Resumo Mensal"
        aplicar_formatacao(ws, "A1:E1", font_size=12, bold=True, fill_color="B8CCE4")
            #limites na planilha toda

        aplicar_formatacao(ws, "A1:W117", border=True)

            #Coluna A

        aplicar_formatacao(ws, "A1:A117", font_size=12, fill_color="8DB3E2")
        ws["A1"].value = "Resumo Mensal"
        aplicar_formatacao(ws, "A1:F1", font_size=12, bold=True, fill_color="B8CCE4")
            #limites na planilha toda

        aplicar_formatacao(ws, "A1:W117", border=True)

            #Coluna A

        aplicar_formatacao(ws, "A1:A117", font_size=12, fill_color="8DB3E2")




    elif opcao == "Folha de Ponto (Fiscalização) Com AV":

         ws["A4"].value = "Resumo Mensal"
         logo_url = "https://storage.googleapis.com/hostinger-horizons-assets-prod/3f69205a-c8f5-4046-89f2-58ab4070596a/a445b8a745a1d79fa57a195f52fac4b4.png"
         response = requests.get(logo_url)
         img_bytes = io.BytesIO(response.content)
         img = Image(img_bytes)
         img.height = 40
         img.width = 100

         ws.add_image(img, "A1")
    
    
         aplicar_formatacao(ws, "A4:S4", border=True, fill_color="BFBFBF")
         ws.merge_cells("M4:S4")
         aplicar_formatacao(ws, "C6:S6", border=True, fill_color="17365D", font_color="FFFFFF")
         aplicar_formatacao(ws, "A7:B7", border=True, bold=True, fill_color="D8D8D8")
         aplicar_formatacao(ws, "A7:B7", border=True, fill_color="D8D8D8")
         aplicar_formatacao(ws, "C7:D7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "F7:G7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "I7:J7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "L7:M7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "O7:P7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "Q7:S7", border=True, fill_color="B8CCE4", font_color="FFFFFF")
         aplicar_formatacao(ws, "E7:E123", fill_color="B8CCE4")
         aplicar_formatacao(ws, "H7:H123", fill_color="B8CCE4")
         aplicar_formatacao(ws, "K7:K123", fill_color="B8CCE4")
         aplicar_formatacao(ws, "N7:N123", fill_color="B8CCE4")
         aplicar_formatacao(ws, "A8:S123", border=True)

    elif opcao == "Folha de Ponto (Fiscalização) Sem AV":
         
         ws["A4"].value = "Resumo Mensal"
         logo_url = "https://storage.googleapis.com/hostinger-horizons-assets-prod/3f69205a-c8f5-4046-89f2-58ab4070596a/a445b8a745a1d79fa57a195f52fac4b4.png"
         response = requests.get(logo_url)
         img_bytes = io.BytesIO(response.content)
         img = Image(img_bytes)
         img.height = 40
         img.width = 100

         ws.add_image(img, "A1")
    
    
         aplicar_formatacao(ws, "A4:S4", border=True, fill_color="BFBFBF")
         ws.merge_cells("M4:S4")
         aplicar_formatacao(ws, "C6:L6", border=True, fill_color="17365D", font_color="FFFFFF")
         aplicar_formatacao(ws, "A7:B7", border=True, bold=True, fill_color="D8D8D8")
         aplicar_formatacao(ws, "A7:B7", border=True, fill_color="D8D8D8")
         aplicar_formatacao(ws, "C7:D7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "E7:F7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "G7:H7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "I7:J7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "K7:L7", border=True, fill_color="BFBFBF", font_color="FFFFFF")
         aplicar_formatacao(ws, "A8:L123", border=True)


    # === Salvar no buffer para download ===
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        label=f"Baixar Planilha {opcao}",
        data=output,
        file_name=f"{opcao.lower().replace(' ', '_')}_formatado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
